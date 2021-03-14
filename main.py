import scraper
import requests
from openpyxl import load_workbook

ROW_COUNT = 1623
FIRST_ROW = 1
A_COLUMN = 1
B_COLUMN = 2


def download_images():
    workbook = load_workbook(filename="data/find-images.xlsx")
    sheet = workbook.active

    for row in range(2, ROW_COUNT):
        image_code = sheet.cell(row, 1).value
        image_url = scraper.get_image(image_code)

        image_extension = image_url.split('.')[-1]
        image_name = f"images/{image_code}.{image_extension}"

        with open(image_name, 'wb') as file:
            file.write(requests.get(image_url).content)
            print(f"Image_downloaded: {image_name}")


def write_image_url():
    workbook = load_workbook(filename="data/find-images.xlsx")
    sheet = workbook.active

    sheet.cell(FIRST_ROW, B_COLUMN, value="Product Image URL")
    for row in range(2, ROW_COUNT):
        image_code = sheet.cell(row, A_COLUMN).value
        image_url = scraper.get_image(image_code)
        sheet.cell(row, B_COLUMN, value=image_url)
        print(image_url)

    workbook.save("data/with_image_urls.xlsx")


if __name__ == '__main__':
    write_image_url()
